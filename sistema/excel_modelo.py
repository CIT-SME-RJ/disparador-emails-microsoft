import shutil

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

from sistema.config import PASTA_PLANILHAS


def obter_caminho_copia_modelo_disponivel():
    contador = 2

    while True:
        caminho_copia = PASTA_PLANILHAS / f"Planilha_Modelo ({contador}).xlsx"

        if not caminho_copia.exists():
            return caminho_copia

        contador += 1


def criar_arquivo_planilha_modelo(caminho_destino):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Dados"

    colunas_modelo = [
        "Enviar",
        "Email_Destino",
        "Nome",
        "Setor",
        "Anexo_Personalizado"
    ]

    dados_modelo = [
        ["Sim", "exemplo@dominio.com", "Maria Exemplo", "Setor Exemplo", "arquivo_exemplo.pdf"],
        ["", "", "", "", ""]
    ]

    worksheet.append(colunas_modelo)

    for linha in dados_modelo:
        worksheet.append(linha)

    preenchimento_obrigatorio = PatternFill(
        start_color="FFD966",
        end_color="FFD966",
        fill_type="solid"
    )

    preenchimento_opcional = PatternFill(
        start_color="D9EAD3",
        end_color="D9EAD3",
        fill_type="solid"
    )

    fonte_cabecalho = Font(bold=True, color="000000")

    alinhamento_cabecalho = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    borda_fina = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7")
    )

    comentarios = {
        "Enviar": "Obrigatório. Use Sim, S, X, 1, True, Enviar ou OK para marcar a linha para envio.",
        "Email_Destino": "Obrigatório. Informe o e-mail do destinatário.",
        "Nome": "Opcional. Pode ser usado no corpo do e-mail com a tag {Nome}.",
        "Setor": "Opcional. Pode ser usado no corpo do e-mail com a tag {Setor}.",
        "Anexo_Personalizado": "Opcional. Informe o nome do arquivo que está na pasta Anexos_Personalizados. Para vários anexos, separe por ponto e vírgula."
    }

    colunas_obrigatorias = {"Enviar", "Email_Destino"}

    for celula in worksheet[1]:
        nome_coluna = celula.value

        celula.font = fonte_cabecalho
        celula.alignment = alinhamento_cabecalho
        celula.border = borda_fina

        if nome_coluna in colunas_obrigatorias:
            celula.fill = preenchimento_obrigatorio
        else:
            celula.fill = preenchimento_opcional

        texto_comentario = comentarios.get(nome_coluna)

        if texto_comentario:
            celula.comment = Comment(texto_comentario, "Sistema")

    for linha in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for celula in linha:
            celula.border = borda_fina
            celula.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:E{worksheet.max_row}"

    larguras = {
        "A": 14,
        "B": 32,
        "C": 24,
        "D": 24,
        "E": 38
    }

    for coluna, largura in larguras.items():
        worksheet.column_dimensions[coluna].width = largura

    worksheet["G1"] = "Legenda"
    worksheet["G1"].font = Font(bold=True)

    worksheet["G2"] = "Amarelo = coluna obrigatória"
    worksheet["G2"].fill = preenchimento_obrigatorio

    worksheet["G3"] = "Verde = coluna opcional"
    worksheet["G3"].fill = preenchimento_opcional

    worksheet.column_dimensions["G"].width = 34

    workbook.save(caminho_destino)


def criar_planilha_modelo():
    caminho_modelo = PASTA_PLANILHAS / "Planilha_Modelo.xlsx"

    if caminho_modelo.exists():
        caminho_copia = obter_caminho_copia_modelo_disponivel()
        shutil.copy2(caminho_modelo, caminho_copia)

        return {
            "tipo": "copia",
            "caminho": caminho_copia,
            "mensagem": f"Foi criada uma cópia da planilha modelo: {caminho_copia.name}"
        }

    criar_arquivo_planilha_modelo(caminho_modelo)

    return {
        "tipo": "nova",
        "caminho": caminho_modelo,
        "mensagem": f"Foi criada a planilha modelo: {caminho_modelo.name}"
    }